from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

window = Tk()
window.title ("Bank Details")
window.geometry("700x450")
window.resizable(False,False)
window.configure(bg="#326273")


file = pathlib.Path("Backened_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet= file.active
    sheet["A1"]="Full Name"
    sheet["B1"]="AccountNumber"
    sheet["C1"]="ContactNumber"
    sheet["D1"]="Gender"
    sheet["E1"]="NIN"
    sheet["F1"]="Address"

    file.save(r"Backened_data.xlsx")


def submit():
    name = namevalue.get()
    contact = contactvalue.get()
    account = accountvalue.get()
    nin = ninvalue.get()
    address = addressentry.get(1.0, END)
    gender = gender_combobox.get()

    file = openpyxl.load_workbook("Backened_data.xlsx")
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+1, value = name)
    sheet.cell(column=2, row=sheet.max_row, value = contact)
    sheet.cell(column=3, row=sheet.max_row, value = account)
    sheet.cell(column=4, row=sheet.max_row, value = nin)
    sheet.cell(column=5, row=sheet.max_row, value = address)
    sheet.cell(column=6, row=sheet.max_row, value = gender)

    file.save(r"Backened_data.xlsx")

    messagebox.showinfo("info","Details Added")

    namevalue.set("")
    accountvalue.set("")
    ninvalue.set("")
    contactvalue.set("")
    addressentry.delete(1.0,END)








def clear():
    namevalue.set("")
    contactvalue.set("")
    accountvalue.set("")
    ninvalue.set("")
    addressentry.delete(1.0, END)






labelhead1 = Label(window, text="Please fill out Your Bank Details", font=("Arial",13), fg= "black" )
labelhead1.place(x=20,y=20)

namelabel = Label(window, text="Name", font=23, bg="#326273",fg="black" )
namelabel.place(x=50, y=100)
contlabel = Label(window, text="Contact No", font=23, bg="#326273",fg="black" )
contlabel.place(x=50, y=150)
accountlabel = Label(window, text="Account No", font=23, bg="#326273",fg="black" )
accountlabel.place(x=50, y=200)
NiNlabel = Label(window, text="Gender", font=23, bg="#326273",fg="black" )
NiNlabel.place(x=370, y=200)
Genderlabel = Label(window, text="NIN", font=23, bg="#326273",fg="black" )
Genderlabel.place(x=50, y=250)
addresslabel = Label(window, text="Address", font=23, bg="#326273", fg="black")
addresslabel.place(x=50, y=300)


namevalue = StringVar()
contactvalue = StringVar()
accountvalue = StringVar()
ninvalue = StringVar()


nameentry = Entry(window, textvariable=namevalue, width=45, bd=2, font=20)
nameentry.place(x=200,y=100)
contentry = Entry(window, textvariable=contactvalue, width=45, bd=2, font=20)
contentry.place(x=200,y=150)
accountentry = Entry(window, textvariable=accountvalue, width=15, bd=2, font=20)
accountentry.place(x=200,y=200)
NINentry = Entry(window, textvariable=ninvalue, width=45, bd=2, font=20)
NINentry.place(x=200,y=250)
addressentry = Text( window, width=50, height=4, bd=4)
addressentry.place(x=200, y=300)


gender_combobox = Combobox(window, values=["male","female"], font="Arial",state="r", width=14)
gender_combobox.place(x=440, y=200)


Button(window, text="Submit", bg="#326273", fg="white", width=15,
 height=2,activebackground="#326273",command=submit).place(x=200 ,y= 400)
Button(window, text="Clear", bg="#326273", fg="white", width=15,
 height=2,activebackground="#326273",command=clear).place(x=340 ,y= 400)
Button(window, text="Exit", bg="#326273", fg="white", width=15,
 height=2,activebackground="#326273", command=lambda: window.destroy()).place(x=480 ,y= 400)




window.mainloop()